import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog
import pandas as pd
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")
import winsound
import os
import pickle
from openpyxl import load_workbook
from openpyxl.styles import Font
from PIL import Image, ImageTk
class WarehousePacker:
    def __init__(self, root):
        self.root = root
        self.root.title("Warehouse Packer")
        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Arial", 16))
        style.configure("Treeview", font=("Arial", 20))  # Если нужно и сами строки
        style.configure("Treeview", rowheight=36)
        self.row_height = 36
        self.img = ImageTk.PhotoImage(Image.new('RGBA', (1, self.row_height), (255, 255, 255, 0)))

        # Data structures
        self.data = None         
        self.gtin_map = None         
        self.packages = {}         
        self.current_box = None

        # Persistent GTIN mapping file
        self.mapping_file = os.path.expanduser('~/.warehouse_packer_gtin.pkl')
        self._load_mapping_disk()

        # Build UI
        main_frame = tk.Frame(root)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Left panel: boxes
        box_frame = tk.Frame(main_frame, bd=2, relief=tk.GROOVE)
        box_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)
        tk.Label(box_frame, text="Коробки:").pack(pady=5)
        self.box_listbox = tk.Listbox(box_frame, height=15)
        self.box_listbox.pack(fill=tk.Y, padx=5)
        self.box_listbox.bind('<<ListboxSelect>>', self.on_box_select)
        btn_frame = tk.Frame(box_frame)
        btn_frame.pack(pady=5)
        for text, cmd in [("Добавить", self.add_box), ("Переименовать", self.rename_box), ("Удалить", self.delete_box)]:
            tk.Button(btn_frame, text=text, command=cmd).pack(side=tk.LEFT, padx=5)

        # Right panel: controls and tree
        right_frame = tk.Frame(main_frame)
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        toolbar = tk.Frame(right_frame)
        toolbar.pack(fill=tk.X, pady=5)
        actions = [
            ("Загрузить лист", self.load_sheet),
            ("Загрузить GTIN", self.load_gtin_map),
            ("Скачать шаблон", self.download_template),
            ("Экспорт", self.export),
            ("Отгрузка WB", self.ship_wb),
            ("Отгрузка Ozon", self.ship_ozon)
        ]
        for text, cmd in actions:
            tk.Button(toolbar, text=text, command=cmd).pack(side=tk.LEFT, padx=5)

        scan_frame = tk.Frame(right_frame)
        scan_frame.pack(fill=tk.X, pady=5)
        tk.Label(scan_frame, text="Сканер (GTIN):").pack(side=tk.LEFT, padx=5)
        self.scan_entry = tk.Entry(scan_frame)
        self.scan_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        self.scan_entry.bind('<Return>', self.process_scan)
        self.scan_entry.focus_set()

        # Tree with dynamic headings
        self.tree = ttk.Treeview(right_frame, columns=("article","scanned","remaining"), show='headings')
        self.tree.heading('article', text='Артикул')
        self.tree.heading('scanned', text='В коробке')
        self.tree.heading('remaining', text='Осталось')
        self.tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.tree.bind('<Double-1>', self.on_tree_double_click)

        # Label: total remaining
        self.remaining_label = tk.Label(right_frame, text="")
        self.remaining_label.pack(pady=3)

    def _load_mapping_disk(self):
        if os.path.exists(self.mapping_file):
            try:
                with open(self.mapping_file, 'rb') as f:
                    self.gtin_map = pickle.load(f)
            except:
                self.gtin_map = None
        else:
            self.gtin_map = None

    def _save_mapping_disk(self):
        if self.gtin_map is not None:
            with open(self.mapping_file, 'wb') as f:
                pickle.dump(self.gtin_map, f)

    def load_sheet(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files","*.xls *.xlsx")])
        if not path: return
        try:
            df = pd.read_excel(path)
            cols = {c.lower(): c for c in df.columns}
            if 'артикул' in cols and 'количество' in cols:
                sub = df[[cols['артикул'], cols['количество']]].copy()
                sub.columns = ['article','quantity']
                sub['quantity'] = sub['quantity'].astype(int)
                sub.set_index('article', inplace=True)
                self.data = sub
                self.data.sort_index(inplace=True)
            else:
                col0, col1 = df.columns[:2]
                df = df.astype({col0: str, col1: int})
                df.columns = ['article','quantity']
                df.set_index('article', inplace=True)
                self.data = df
                self.data.sort_index(inplace=True)

            messagebox.showinfo("Готово", f"Загружено {len(self.data)} позиций.")
            self.packages.clear(); self.current_box=None
            self.box_listbox.delete(0, tk.END)
            self.tree.delete(*self.tree.get_children())
            self.refresh_tree()
        except Exception as e:
            winsound.Beep(1000,200)
            messagebox.showerror("Ошибка", f"Не удалось загрузить лист:\n{e}")

    def load_gtin_map(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files","*.xls *.xlsx")])
        if not path: return
        try:
            df = pd.read_excel(path, dtype=str)
            df.columns = ['gtin','article']
            df.set_index('gtin', inplace=True)
            self.gtin_map = df['article']
            self._save_mapping_disk()
            messagebox.showinfo("Готово", f"Загружено {len(self.gtin_map)} GTIN-сопоставлений.")
        except Exception as e:
            winsound.Beep(1000,200)
            messagebox.showerror("Ошибка", f"Не удалось загрузить GTIN-таблицу:\n{e}")

    def download_template(self):
        df = pd.DataFrame(columns=['Артикул','Количество'])
        path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel','*.xlsx')])
        if not path: return
        try:
            df.to_excel(path, index=False)
            messagebox.showinfo("Готово", f"Шаблон сохранён в {os.path.basename(path)}")
        except Exception as e:
            winsound.Beep(1000,200)
            messagebox.showerror("Ошибка", f"Не удалось сохранить шаблон:\n{e}")

    def add_box(self):
        if self.data is None:
            messagebox.showwarning("Внимание","Сначала загрузите лист.")
            return
        name = simpledialog.askstring("Имя коробки","Введите имя новой коробки:")
        if not name or name in self.packages: return
        self.packages[name] = {art:0 for art in self.data.index}
        self.box_listbox.insert(tk.END,name)
        self.box_listbox.selection_clear(0, tk.END)
        self.box_listbox.selection_set(tk.END)
        self.on_box_select()

    def rename_box(self):
        sel = self.box_listbox.curselection()
        if not sel: return
        old = self.box_listbox.get(sel)
        new = simpledialog.askstring("Переименовать","Новое имя:", initialvalue=old)
        if not new or new in self.packages: return
        self.packages[new] = self.packages.pop(old)
        self.box_listbox.delete(sel); self.box_listbox.insert(sel,new); self.box_listbox.selection_set(sel)
        self.on_box_select()

    def delete_box(self):
        sel = self.box_listbox.curselection()
        if not sel: return
        name = self.box_listbox.get(sel)
        if messagebox.askyesno("Удалить", f"Удалить '{name}'?"):
            self.packages.pop(name,None)
            self.box_listbox.delete(sel)
            self.current_box=None; self.tree.delete(*self.tree.get_children())
            self.refresh_tree()

    def on_box_select(self,event=None):
        sel = self.box_listbox.curselection()
        if not sel: return
        self.current_box = self.box_listbox.get(sel)
        self.refresh_tree()

    def total_scanned(self,article):
        return sum(b.get(article,0) for b in self.packages.values())

    def process_scan(self, event):
        gtin = self.scan_entry.get().strip()
        self.scan_entry.delete(0, tk.END)
        if self.data is None or self.current_box is None or self.gtin_map is None:
            winsound.Beep(1000,200)
            messagebox.showwarning("Внимание","Загрузите данные и выберите коробку.")
            self.scan_entry.focus_set()
            return
        if gtin not in self.gtin_map.index:
            winsound.Beep(1000,200)
            messagebox.showwarning("Не найден GTIN", f"GTIN {gtin} отсутствует.")
            return
        article = self.gtin_map.at[gtin]
        if article not in self.data.index:
            winsound.Beep(1000,200)
            messagebox.showerror("Ошибка данных", f"Артикул {article} не найден в листе.")
            return
        allowed = self.data.at[article,'quantity']
        used = self.total_scanned(article)
        remaining = allowed - used
        if remaining <= 0:
            winsound.Beep(1000,200)
            messagebox.showerror("Превышено", f"Доступно {allowed}, использовано {used}")
            return

        # Record successful scan and play success sound
        self.packages[self.current_box][article] += 1
        winsound.PlaySound('SystemAsterisk', winsound.SND_ALIAS | winsound.SND_ASYNC)
        self.refresh_tree()
        self.scan_entry.focus_set()

    def on_tree_double_click(self,event):
        item=self.tree.identify_row(event.y); col=self.tree.identify_column(event.x)
        if not item or col!='#2': return
        art, scanned, _= self.tree.item(item,'values')
        new_val=simpledialog.askinteger("Редактировать","Новое количество:", initialvalue=int(scanned), minvalue=0)
        if new_val is None: return
        allowed=self.data.at[art,'quantity']
        other=self.total_scanned(art)-int(scanned)
        if new_val+other>allowed:
            winsound.Beep(1000,200)
            messagebox.showerror("Превышено", f"Всего доступно {allowed}, в других коробках {other}")
            return
        self.packages[self.current_box][art]=new_val
        self.refresh_tree()

    def refresh_tree(self):
        # Update headings with totals
        total_articles = len(self.data) if self.data is not None else 0
        if self.current_box is not None:
            total_scanned = sum(self.packages[self.current_box].values())
        else:
            total_scanned = 0
        total_remaining = sum(self.data.at[art,'quantity'] - self.total_scanned(art) for art in self.data.index) if self.data is not None else 0

        self.tree.heading('article', text=f'Артикул ({total_articles})')
        self.tree.heading('scanned', text=f'В коробке ({total_scanned})')
        self.tree.heading('remaining', text=f'Осталось ({total_remaining})')

        # Refresh rows
        self.tree.delete(*self.tree.get_children())
        select_next = None
        if self.data is not None and self.current_box is not None:
            for idx, (art, qty) in enumerate(self.data['quantity'].items()):
                scanned = self.packages[self.current_box].get(art, 0)
                rem = qty - self.total_scanned(art)
                iid = self.tree.insert('', tk.END, values=(art, scanned, rem))
                if select_next is None and rem > 0:
                    select_next = iid
        # выделить строку с первым незаполненным
        if select_next is not None:
            self.tree.selection_set(select_next)
            self.tree.focus(select_next)
            self.tree.see(select_next)

        # Update label with total remaining for ALL boxes
        self.remaining_label.config(text=f"Всего осталось распределить: {total_remaining}")

    # ... остальные функции (экспорт, ship_wb, ship_ozon) не менялись ...

    def export(self):
        rows=[]
        for box,items in self.packages.items():
            for art,cnt in items.items():
                if cnt>0:
                    rows.append({'Артикул товара':art,'Кол-во товаров':cnt,'Коробка':box})
        if not rows:
            winsound.Beep(1000,200); messagebox.showwarning("Пусто","Нет данных для экспорта."); return
        df=pd.DataFrame(rows)
        path=filedialog.asksaveasfilename(defaultextension='.xlsx',filetypes=[('Excel','*.xlsx')])
        if not path: return
        try:
            df.to_excel(path,index=False)
            messagebox.showinfo("Готово",f"Сохранено в {os.path.basename(path)}")
        except Exception as e:
            winsound.Beep(1000,200); messagebox.showerror("Ошибка",f"Не удалось сохранить:\n{e}")

    def ship_wb(self):
        if not self.packages:
            messagebox.showwarning("Пусто", "Нет данных для отгрузки WB.")
            return
        tpl_path = filedialog.askopenfilename(title="Загрузить шаблон WB", filetypes=[('Excel','*.xlsx')])
        if not tpl_path: return
        try:
            tpl = pd.read_excel(tpl_path, dtype=str)
            if 'ШК короба' not in tpl.columns or 'Срок годности' not in tpl.columns:
                raise ValueError('Шаблон должен содержать колонки "ШК короба" и "Срок годности"')
        except Exception as e:
            winsound.Beep(1000,200)
            messagebox.showerror("Ошибка шаблона", str(e)); return
        boxes = list(self.packages.keys())
        # if len(tpl) != len(boxes):
        #     winsound.Beep(1000,200)
        #     messagebox.showerror("Несоответствие", "Количество строк шаблона не равно количеству коробок"); return
        article_to_gtin = {art:gt for gt,art in self.gtin_map.items()} if self.gtin_map is not None else {}
        out_rows = []
        for idx, box in enumerate(boxes):
            box_code = tpl.at[idx, 'ШК короба']
            shelf_life = tpl.at[idx, 'Срок годности']
            for art, cnt in self.packages[box].items():
                if cnt > 0:
                    barcode = article_to_gtin.get(art, art)
                    out_rows.append({
                        'Баркод товара': barcode,
                        'Кол-во товаров': cnt,
                        'ШК короба': box_code,
                        'Срок годности': shelf_life
                    })
        df_out = pd.DataFrame(out_rows)
        save_path = filedialog.asksaveasfilename(defaultextension='.xlsx', title="Сохранить отгрузку WB", filetypes=[('Excel','*.xlsx')])
        if not save_path: return
        try:
            # write without bold or borders
            df_out.to_excel(save_path, index=False)
            wb = load_workbook(save_path)
            ws = wb.active
            # remove bold from header and clear borders
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.font = Font(bold=False)
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = None
            wb.save(save_path)
            messagebox.showinfo("Готово", f"WB отгрузка сохранена в {os.path.basename(save_path)}")
        except Exception as e:
            winsound.Beep(1000,200)
            messagebox.showerror("Ошибка", f"Не удалось сохранить WB файл:\n{e}")

    def ship_ozon(self):
            if not self.packages:
                messagebox.showwarning("Пусто", "Нет данных для отгрузки Ozon.")
                return
            tpl_path = filedialog.askopenfilename(title="Загрузить шаблон Ozon", filetypes=[('Excel','*.xlsx')])
            if not tpl_path: return
            try:
                tpl = pd.read_excel(tpl_path, dtype=str)
                required = ['ШК товара','Артикул товара','Кол-во товаров','Зона размещения','ШК ГМ','Тип ГМ (не обязательно)','Срок годности ДО в формате YYYY-MM-DD (не более 1 СГ на 1 SKU в 1 ГМ)']
                missing = [col for col in required if col not in tpl.columns]
                if missing:
                    raise ValueError(f"Шаблон должен содержать колонки: {', '.join(missing)}")
            except Exception as e:
                winsound.Beep(1000,200)
                messagebox.showerror("Ошибка шаблона", str(e))
                return
            boxes = list(self.packages.keys())

            article_to_gtin = {art:gt for gt,art in self.gtin_map.items()} if self.gtin_map is not None else {}
            out_rows = []
            for idx, box in enumerate(boxes):
                row = tpl.iloc[idx]
                sku = row['ШК товара']
                zone = row['Зона размещения']
                gm_code = row['ШК ГМ']
                gm_type = row.get('Тип ГМ', '')
                shelf = row['Срок годности ДО в формате YYYY-MM-DD (не более 1 СГ на 1 SKU в 1 ГМ)']
                for art, cnt in self.packages[box].items():
                    if cnt > 0:
                        barcode = article_to_gtin.get(art, art)
                        out_rows.append({
                            'ШК товара': barcode,
                            'Артикул товара': art,
                            'Кол-во товаров': cnt,
                            'Зона размещения': zone,
                            'ШК ГМ': gm_code,
                            'Тип ГМ (не обязательно)': gm_type,
                            'Срок годности ДО в формате YYYY-MM-DD (не более 1 СГ на 1 SKU в 1 ГМ)': shelf
                        })
            df_out = pd.DataFrame(out_rows)
            save_path = filedialog.asksaveasfilename(defaultextension='.xlsx', title="Сохранить отгрузку Ozon", filetypes=[('Excel','*.xlsx')])
            if not save_path: return
            try:
                df_out.to_excel(save_path, index=False)
                messagebox.showinfo("Готово", f"Ozon отгрузка сохранена в {os.path.basename(save_path)}")
            except Exception as e:
                winsound.Beep(1000,200)
                messagebox.showerror("Ошибка", f"Не удалось сохранить Ozon файл:\n{e}")

if __name__=='__main__':
    root=tk.Tk()
    WarehousePacker(root)
    root.mainloop()
